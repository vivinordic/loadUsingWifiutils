#-------------------------------------------------------------------------------
# Name:        aci
# Purpose:     This script will calculate the aci value for the specified datarate.
# Author:      vivi
# Created:     20-09-2021
# Copyright:   Nordic Semiconductors
#-------------------------------------------------------------------------------

from rs import *
from iqxel import *
from common_utils import *
from harness import *
from input import *
import numpy
from openpyxl import *

op_file_path=""
rw_num=1
op_file_path=""
sensitiviy_dr = {}

#Loop iterating for the range of amplitudes mentioned


def amplitude_loop(dr,ch):
    start_amp_list=[]
    global row_num
    global per_evenrow_dict
    per_evenrow_dict={}
    global per_oddrow_dict
    per_oddrow_dict={}
    global even_xaxis
    global odd_xaxis
    global data
    global row_num_dr
    row_num_dr=2
##    worksheet.write_row('A1',['Standard','Channel','DataRate','LOW_AMPLITUDE','POP_AMPLITUDE','ED_CNT','CRC32_PASS_CNT','CRC32_FAIL_CNT','OFDM_CORR_PASS','DSSS_CORR_PASS','L1_CORR_FAIL_CNT','LSIG_FAIL_CNT','HTSIG_FAIL_CNT','VHTSIGA_FAIL_CNT','VHTSIGB_FAIL_CNT','HESIGA_FAIL_CNT','HESIGB_FAIL_CNT','DSSS_SFD_FAIL_CNT','DSSS_HDR_FAIL_CNT','POP_CNT','MID_PACKET_CNT','UNSUPPORTED_CNT','OTHER_STA_CNT','SpatialReuse_CNT','PER %'],bold1)
##    worksheet_dr.write_row('A1',['Standard','Channel','DataRate','LOW_AMPLITUDE','POP_AMPLITUDE','ED_CNT','CRC32_PASS_CNT','CRC32_FAIL_CNT','OFDM_CORR_PASS','DSSS_CORR_PASS','L1_CORR_FAIL_CNT','LSIG_FAIL_CNT','HTSIG_FAIL_CNT','VHTSIGA_FAIL_CNT','VHTSIGB_FAIL_CNT','HESIGA_FAIL_CNT','HESIGB_FAIL_CNT','DSSS_SFD_FAIL_CNT','DSSS_HDR_FAIL_CNT','POP_CNT','MID_PACKET_CNT','UNSUPPORTED_CNT','OTHER_STA_CNT','SpatialReuse_CNT','PER %'],bold)

    axis_num=1
    flg=0
    a=1
    global limit_trail
    global flag
    flag=0
    limit_trail=1
    res=dut.dut_read(1500)
    for amp in ampl_dict[dr]:
        data=[]
        data.append(standard)
        data.append(int(ch))
        try:
            data.append(int(dr))
        except:
            data.append(dr)
        data.append(firstPktPow)
        data.append(float(amp))
        data.append(step_size)
        if(flg==0):
            if(ampl_dict[dr].index(amp)==0):
                equipment.set_aci_amplitude(streams,str(int(amp)+1))
            equipment.set_aci_amplitude(streams,str(amp+cable_loss+splitter_loss))
            data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
            #if((data[4] < 200 ) and ((ampl_dict[dr].index(amp)==0) or (ampl_dict[dr].index(amp)==1))):
            if(data[6] < 800 ):# and ((ampl_dict[dr].index(amp)==0) or (ampl_dict[dr].index(amp)==1))):
                time.sleep(1)
                dut.set_dut_channel(int(ch))
                for i in range(num_params_to_pop):
                    data.pop()
                data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
            if((data[6] > 3000) and ('IQXEL' in tester)):
                time.sleep(2)
                for i in range(num_params_to_pop):
                    data.pop()
                data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
            print data
            if(data[6]=='LOCKUP'):
                for i in range(num_params_to_pop):
                    data.pop()
                if(dut.check_dut_stuck_state()=='alive'):
                    dut.dut_reboot()
                else:
                    print 'DUT is stuck'
                    controlPowerSwitch(on_ports_list='8',off_ports_list='8')
                time.sleep(reboot_time)
                dut.init_dut(standard=standard,streams=streams,channel=int(ch),bw=bw,release=release,chain_sel=chain_sel)
                dut.set_dut_production_mode_settings(standard=standard,ch=int(ch),bw=bw)
                dut.set_dut_channel(int(ch))
                data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                print data
            worksheet.append(data)


def pop_analysis(DUT_TestConfigParams,dut1):
    #controlPowerSwitch(on_ports_list='8',off_ports_list='8')
    global ftime
    global equipment
    global dut
    global modulation
    global row_num
    global standard
    global streams_chain
    global streams
    global freq
    global bw
    global gi
    global stbc
    global coding
    global data_rate
    global cable_loss
    global greenfield_mode
    global channel
    global op_file_path
    global preamble
    global workbook
    global workbook_dr
    global worksheet
    global worksheet_dr
    global bold
    global bold1
    global firstPktPow
    global ampl_dict
    global aci_amp_dict
    global step_size
    greenfield_mode=stbc=preamble=gi=coding=preamble=''
    global ftime
    global tester
    ftime=open('total_time','a')
    ftime.write(time.strftime("%H-%M-%S")+"\n")
    ampl_dict={}
    actual_ampl_dict={}
    aci_amp_dict={}
    row_num=2
    dut = dut1

    standard = DUT_TestConfigParams.standard
    streams=DUT_TestConfigParams.streams
    chain_sel=DUT_TestConfigParams.chain_sel
    streams_chain = streams+'_'+str(chain_sel)
    bw=DUT_TestConfigParams.bw
    data_rate=DUT_TestConfigParams.data_rate
    dutModel=DUT_TestConfigParams.dutModel
    actual_start_amplt=DUT_TestConfigParams.start_amplt
    actual_end_amplt=DUT_TestConfigParams.end_amplt
    step_size=DUT_TestConfigParams.step_size
    channel=DUT_TestConfigParams.channel
    payload=DUT_TestConfigParams.payload
    if(standard=='11b'):
        preamble=DUT_TestConfigParams.preamble

    else:
        stbc=DUT_TestConfigParams.stbc
        gi=DUT_TestConfigParams.gi
        coding=DUT_TestConfigParams.coding
        if(standard=='11n'):
            greenfield_mode=DUT_TestConfigParams.greenfield_mode


    if(int(channel.split(',')[-1])<30):
        freq='2.4GHz'
    else:
        freq='5GHz'
    op_file_path=BuildResultsPath(DUT_TestConfigParams)
    time.sleep(0.5)
    if(standard=='11b'):
        modulation='DSSS'
    else:
        modulation='OFDM'
    ftime.write(time.strftime("%H-%M-%S")+"\n")
    DebugPrint(create='1')
    DebugPrint('Started')

    for dr in data_rate.split(','):
        amplitude_list=[]
        if(1 <= int(channel) <= 14):
            cable_loss_band='24G_BAND'
        if(36 <= int(channel) <= 52):
            cable_loss_band='5G_BAND1'
        elif(53 <= int(channel) <= 108):
            cable_loss_band='5G_BAND2'
        elif(109 <= int(channel) <= 132):
            cable_loss_band='5G_BAND3'
        elif(133 <= int(channel) <= 165):
            cable_loss_band='5G_BAND4'
        cable_loss=cable_loss_dict['1x1'][cable_loss_band]
        DUT_TestConfigParams.data_rate = dr
        start_amplt = actual_start_amplt
        firstPktPow = actual_end_amplt
        steps_list=list(numpy.arange(start_amplt,start_amplt+1,2))
        #steps_list.reverse()
        for j in steps_list:
            amplitude_list.append(j)
        #if 0 not in amplitude_list:
        #    amplitude_list=[0]+amplitude_list
        print amplitude_list
        ampl_dict[dr]=amplitude_list
##        actual_amplitude_list=[]
##        steps_list=list(numpy.arange(actual_end_amplt,actual_start_amplt+1,step_size))
##        #steps_list.reverse()
##        for j in steps_list:
##            actual_amplitude_list.append(j)
##        actual_ampl_dict[dr]=actual_amplitude_list

##    dut=eval(dutModel)(com_port)
    tester = DUT_TestConfigParams.vsg
    equipment=eval(tester.split('_')[0])(equip_ip = DUT_TestConfigParams.equip_ip)
    DebugPrint('Trying to initialize dut in PER Main')
    if(dut.init_dut(standard=standard,streams=streams,channel=int(channel.split(',')[0]),bw=bw,release=release,chain_sel=chain_sel)=='CONNECT_FAIL'):
        print 'CONNECT_FAIL'
        DebugPrint('CONNECT_FAIL in PER Main')
        controlPowerSwitch(on_ports_list='8',off_ports_list='8')
        if(dut.init_dut(standard=standard,streams=streams,channel=int(channel.split(',')[0]),bw=bw,release=release,chain_sel=chain_sel)=='CONNECT_FAIL'):
            print 'Switching to Serial Access'
            DebugPrint('Switching to Serial Access')
            dutModel=dutModel.split('_')[0]
            dut=eval(dutModel)(com_port)
            equipment=eval(tester.split('_')[0])(tester,DUT_TestConfigParams.equip_ip)
            dut.init_dut(standard=standard,streams=streams,channel=int(channel.split(',')[0]),bw=bw,release=release,chain_sel=chain_sel)

    dut.dut_down_up(action='up_down')
    DUT_functions.startTestcase()

    try:
        equipment.init_vsg_funcs(standard=standard,bw=bw,streams=streams,stbc=stbc,gi=gi,coding=coding,greenfield_mode=greenfield_mode,preamble=preamble,payload=payload,run='pop',chain_sel=str(chain_sel))
    except Exception,e:
        print 'Equipment is not reachable'
        exit(0)
    equipment.set_aci_bandwidth(standard=standard,bw=bw)
    equipment.rf_on_off_aci()
    for ch in channel.split(','):
        if(int(ch)<30):
            freq='2.4GHz'
        else:
            freq='5GHz'
        if(1 <= int(ch) <= 14):
            cable_loss_band='24G_BAND'
        if(36 <= int(ch) <= 52):
            cable_loss_band='5G_BAND1'
        elif(53 <= int(ch) <= 108):
            cable_loss_band='5G_BAND2'
        elif(109 <= int(ch) <= 132):
            cable_loss_band='5G_BAND3'
        elif(133 <= int(ch) <= 165):
            cable_loss_band='5G_BAND4'
        cable_loss_1x1=cable_loss_dict['1x1'][cable_loss_band]
        cable_loss=cable_loss_2x2=cable_loss_1x1
        if(streams=='2x2'):
            cable_loss_2x2=cable_loss_dict['2x2'][cable_loss_band]
            cable_loss=float((cable_loss_1x1+cable_loss_2x2)/2)
        #cable_loss=cable_loss+splitter_loss
        dut.set_dut_production_mode_settings(standard=standard,ch=int(ch),bw=bw)
        dut.set_dut_channel(int(ch))
        equipment.set_macheader()
        equipment.set_payload(standard,payload)
        equipment.apply_vsg(bw=bw,chn=int(ch),streams=streams)
        equipment.set_aci_macheader()
        equipment.set_aci_payload(standard,payload)
        equipment.apply_vsg(bw=bw,chn=int(ch),streams=streams,chain_sel=2)
        equipment.set_aci_datarate(modulation,standard,dr)
        equipment.set_aci_idleinterval(1000)

        consld_fname=os.path.join(op_file_path.split(standard)[0],'POP_Consolidated_results.xlsx')
##        workbook = xlsxwriter.Workbook(consld_fname)
##        worksheet = workbook.add_worksheet()
##        bold1 = workbook.add_format({'bold': 1})
        if os.path.exists(consld_fname):
            workbook= load_workbook(consld_fname)
            worksheet = workbook.active
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(['Standard','Channel','DataRate','LOW_AMPLITUDE','POP_AMPLITUDE','PKT_GAP']+rx_report_hdngs)
            bold_font = styles.Font(bold=True)
            for cell in worksheet["1:1"]:
                cell.font = bold_font
        if(ch not in aci_amp_dict.keys()):
            aci_amp_dict[ch]={}
        for dr in data_rate.split(','):

            equipment.set_datarate(modulation,standard,dr)
            equipment.set_aci_datarate(modulation,standard,dr)
            equipment.set_idleinterval(1000,chain_sel=chain_sel)
            equipment.set_aci_idleinterval(1000)
            equipment.generate_waveform(streams=streams)
            equipment.generate_pop_waveform(streams=streams)
            equipment.set_pop_delay(2*int(bw)*step_size)
            ftime.write("AMP LOOP START"+time.strftime("%H-%M-%S")+"\n")
            equipment.set_amplitude(streams,str(firstPktPow+cable_loss+splitter_loss))

            amplitude_loop(dr,ch)
            ftime.write("AMP LOOP END"+time.strftime("%H-%M-%S")+"\n")
    try:
        workbook.save(consld_fname)
    except:
        time.sleep(10)
        try:
            workbook.save(consld_fname)
        except:
            pass

    equipment.rf_on_off(rf_state='off',streams=streams)
    equipment.rf_on_off_aci(rf_state='off',streams=streams)
    equipment.close_vsg()
    dut.dut_close()