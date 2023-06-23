#-------------------------------------------------------------------------------
# Name:        per
# Purpose:     This script will generate PER plots of any datarates/channel.
# Author:      kranthi.kishore
# Created:     14-06-2015
# Copyright:   Imagination Technologies Ltd
#-------------------------------------------------------------------------------

from rs import *
from iqxel import *
from common_utils import *
from harness import *
#from CSUtils import DA
import numpy
import os
from openpyxl import *

rw_num=1
num_params_to_pop = 7

#Loop iterating for the range of amplitudes mentioned
def amplitude_loop(dr,ch,workbook,worksheet,standard,cable_loss,DUT_TestConfigParams,staID):
    """ Triggers VSG to transmit packets to DUT and save RX stats to xlsx  """
    start_amp_list=[]
    global row_num
    global per_evenrow_dict
    per_evenrow_dict={}
    global per_oddrow_dict
    per_oddrow_dict={}
    global even_xaxis
    global odd_xaxis
    global data


    axis_num=1
    flg=0
    a=1

    for amp in ampl_dict[dr]:
        data=[]
        data.append(standard)
        data.append(int(ch))
        data.append(DUT_TestConfigParams.ruAllocation)
        data.append(staID)
        data.append(DUT_TestConfigParams.stbc)
        data.append(DUT_TestConfigParams.coding)
        data.append(DUT_TestConfigParams.doppler)
        data.append(DUT_TestConfigParams.giType)
        data.append(DUT_TestConfigParams.dcm)
        data.append(DUT_TestConfigParams.midamblePeriodicity)
        data.append(DUT_TestConfigParams.heLtfType)
        data.append(DUT_TestConfigParams.sigBmcs)
        data.append(DUT_TestConfigParams.sigBdcm)
        data.append(DUT_TestConfigParams.sigBcompression)
        data.append(DUT_TestConfigParams.comment)
        try:
            data.append(int(dr))
        except:
            data.append(dr)
        data.append(amp-cable_loss)
        if(flg==0):
            equipment.set_amplitude(vsg_streams,str(amp))
            start_per(modulation,dr,ch,vsg_streams,standard,tester,dut,equipment)
            data=data+dut.get_negative_stats()
            DebugPrint(data)
        worksheet.append(data)
        per_amp_dict[dr].append(float(data[8]))
        row_num=row_num+1
        axis_num=axis_num+1
        ftime.write(time.strftime("%H-%M-%S")+"\n")
    #row_num=row_num+1


#if __name__ == "__main__":
def rx_negative_analysis(DUT_TestConfigParams,dut1):
    #controlPowerSwitch(on_ports_list='8',off_ports_list='8')
    global ftime
    global equipment
    global dut
    global vsg_streams
    global modulation
    global row_num
    global standard
    global streams
    global streams_chain
    global freq
    global bw
    global gi
    global stbc
    global coding
    global data_rate
    global cable_loss
    global greenfield_mode
    global channel
    global op_file_path
    global preamble
    global p20_flag
    global ampl_dict
    ampl_dict={}
    greenfield_mode=stbc=preamble=gi=coding=preamble=''
    ftime=open('total_time','a')
    ftime.write(time.strftime("%H-%M-%S")+"\n")
    global per_amp_dict
    global amplitude_list
    global tester
    per_amp_dict = {}
    row_num=2
    dut = dut1

    standard = DUT_TestConfigParams.standard #standard=sys.argv[1].lower()
    #streams_chain=sys.argv[2]
    streams=DUT_TestConfigParams.streams        #streams=streams_chain.split('_')[0]
    chain_sel=DUT_TestConfigParams.chain_sel    #chain_sel=streams_chain.split('_')[-1]
    streams_chain = streams+'_'+str(chain_sel)
    bw=DUT_TestConfigParams.bw                 #bw=sys.argv[3]
    data_rate=DUT_TestConfigParams.data_rate #data_rate=sys.argv[4]
    dutModel=DUT_TestConfigParams.dutModel
    start_amplt=DUT_TestConfigParams.start_amplt       #start_amplt=int(sys.argv[6])
    end_amplt=DUT_TestConfigParams.end_amplt           #end_amplt=int(sys.argv[7])
    step_size=DUT_TestConfigParams.step_size           #step_size=float(sys.argv[8])
    channel=int(DUT_TestConfigParams.channel)               #channel=sys.argv[9]
    payload=DUT_TestConfigParams.payload               #payload=sys.argv[10]
    p20_flag = DUT_TestConfigParams.p20_flag
    if(standard=='11b'):
        preamble=DUT_TestConfigParams.preamble              #preamble=sys.argv[5]
    elif((standard=='11ac')or(standard=='11n')):
        stbc=DUT_TestConfigParams.stbc#stbc=sys.argv[5]
        gi=DUT_TestConfigParams.gi#gi=sys.argv[6]
        coding=DUT_TestConfigParams.coding#coding=sys.argv[7]
        if(standard=='11n'):
            greenfield_mode=DUT_TestConfigParams.greenfield_mode#greenfield_mode=sys.argv[8]
    elif(standard=='11ax'):
        stbc=DUT_TestConfigParams.stbc
        gi=DUT_TestConfigParams.gi
        coding=DUT_TestConfigParams.coding
        giType = str(0.8*(1<<DUT_TestConfigParams.giType))+'us'
        heLtfType = str((1<<DUT_TestConfigParams.heLtfType))+'x'
        tone = ''
        if(DUT_TestConfigParams.format == 5):
            pktFrameFormat='HESU'
        elif(DUT_TestConfigParams.format == 6):
            pktFrameFormat='HEMU'
        elif(DUT_TestConfigParams.format == 7):
            pktFrameFormat='HEERSU'
            if (DUT_TestConfigParams.chBandwidth ==0):
                tone = '242Tone_'
            else:
                tone = '106Tone_'
        elif(DUT_TestConfigParams.format == 8):
            pktFrameFormat='HETB'
        desiredStaIdList = DUT_TestConfigParams.desiredStaIDlist.split(',')


    if(int(channel)<30):
        freq='2.4GHz'
    else:
        freq='5GHz'
    op_file_path=BuildResultsPath(DUT_TestConfigParams)
    time.sleep(0.5)
    if(standard=='11b'):
        modulation='DSSS'
    else:
        modulation='OFDM'
    ftime.write(time.strftime("%H-%M-%S")+"\n")
    DebugPrint(create='1')
    DebugPrint('Started')

    tester = DUT_TestConfigParams.vsg
    equipment=eval(tester.split('_')[0])(tester,DUT_TestConfigParams.equip_ip)
    #dut.serial_access(standard=standard,streams=streams,channel=channel.split(',')[-1],bw=bw,release=release)
    DebugPrint('Trying to initialize dut in PER Main')
    vsg_streams=streams
    if(stbc=='STBC_1'):
        vsg_streams='1x1'
        chain_sel=1
    if(dut.init_dut(standard=standard,streams=streams,channel=channel,bw=bw,release=release,chain_sel=chain_sel)=='CONNECT_FAIL'):
        print 'CONNECT_FAIL'
        DebugPrint('CONNECT_FAIL in PER Main')
        controlPowerSwitch(on_ports_list='8',off_ports_list='8')
        if(dut.init_dut(standard=standard,streams=streams,channel=channel,bw=bw,release=release,chain_sel=chain_sel)=='CONNECT_FAIL'):
            print 'Switching to Serial Access'
            DebugPrint('Switching to Serial Access')
            dutModel=dutModel.split('_')[0]
            dut=eval(dutModel)(com_port)
            equipment=eval(tester.split('_')[0])(tester,DUT_TestConfigParams.equip_ip)
            dut.init_dut(standard=standard,streams=streams,channel=channel,bw=bw,release=release,chain_sel=chain_sel)

    dut.dut_down_up(action='up_down')
    DUT_functions.startTestcase()

    try:
        equipment.init_vsg_funcs(standard=standard,bw=bw,streams=vsg_streams,stbc=stbc,gi=gi,coding=coding,greenfield_mode=greenfield_mode,preamble=preamble,payload=payload,chain_sel=str(chain_sel))
        if(standard == '11ax'):
            equipment.configHE(DUT_TestConfigParams)
            print("configured HE")
    except Exception,e:
        print (e)
        print 'Equipment is not reachable'
        exit(0)
    for dr in data_rate.split(','):
        amplitude_list=[]
        steps_list=list(numpy.arange(end_amplt,start_amplt+1,step_size))
        steps_list.reverse()
        for j in steps_list:
            amplitude_list.append(j)

        ampl_dict[dr]=amplitude_list
        per_amp_dict[dr]=[]
    #for ch in channel.split(','):
    if(int(channel)<30):
        freq='2.4GHz'
    else:
        freq='5GHz'
    if(1 <= int(channel) <= 14):
        cable_loss_band='24G_BAND'
    if(36 <= int(channel) <= 52):
        cable_loss_band='5G_BAND1'
    elif(53 <= int(channel) <= 108):
        cable_loss_band='5G_BAND2'
    elif(109 <= int(channel) <= 132):
        cable_loss_band='5G_BAND3'
    elif(133 <= int(channel) <= 165):
        cable_loss_band='5G_BAND4'
    cable_loss_1x1=cable_loss_dict['1x1'][cable_loss_band]
    cable_loss=cable_loss_2x2=cable_loss_1x1
    if(streams=='2x2'):
        cable_loss_2x2=cable_loss_dict['2x2'][cable_loss_band]
        cable_loss=float((cable_loss_1x1+cable_loss_2x2)/2)
    dut.set_dut_production_mode_settings(standard=standard,ch=channel,bw=bw)
    dut.set_dut_channel(channel,p20_flag)
    equipment.apply_vsg(bw=bw,chn=channel,streams=vsg_streams,p20_flag=p20_flag)
    equipment.set_macheader()
    equipment.set_payload(standard,payload)

    consld_fname=os.path.join(op_file_path.split(standard)[0],'RX_negative_Consolidated_results.xlsx')
    print(consld_fname)
    #workbook = xlsxwriter.Workbook(consld_fname)
    if os.path.exists(consld_fname):
        workbook= load_workbook(consld_fname)
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(['Standard','Channel','RU Allocation','STAID','stbc',\
         'coding', 'doppler', 'giType', 'dcm', 'midamblePeriodicity', 'heLtfType',\
          'sigBMCS', 'sigBDCM', 'sigBCompression', 'Comment','DataRate','AMPLITUDE',\
          'OFDM_CORR_PASS','L1_CORR_FAIL_CNT','LSIG_FAIL_CNT','HTSIG_FAIL_CNT',\
          'VHTSIGA_FAIL_CNT','HESIGA_FAIL_CNT','HESIGB_FAIL_CNT'])

        bold_font = styles.Font(bold=True)
        for cell in worksheet["1:1"]:
            cell.font = bold_font

    for dr in data_rate.split(','):
##        for ii in range(DUT_TestConfigParams.numUser):
        equipment.setDatarateAllUser(DUT_TestConfigParams.numUser, dr)
        equipment.set_datarate(modulation,standard,dr)
        equipment.generate_waveform(streams=vsg_streams)
        ftime.write("AMP LOOP START"+time.strftime("%H-%M-%S")+"\n")
        for staID in desiredStaIdList:
            dut.dut_setStaID(int(staID))
            DUT_functions.setHemuP20Offset(DUT_TestConfigParams.p20_flag)
            amplitude_loop(dr,channel,workbook,worksheet,standard,cable_loss,DUT_TestConfigParams, staID)
        ftime.write("AMP LOOP END"+time.strftime("%H-%M-%S")+"\n")

    try:
        workbook.save(consld_fname)
    except:
        time.sleep(10)
        try:
            workbook.save(consld_fname)
        except:
            pass

    equipment.rf_on_off(rf_state='off',streams=vsg_streams)
    equipment.close_vsg()
    dut.dut_close()