#-------------------------------------------------------------------------------
# Name:        sensitivity
# Purpose:     This script will calculate the sensitivity value for the specified datarate.
# Author:      kranthi.kishore
# Created:     14-06-2015
# Copyright:   Imagination Technologies Ltd
#-------------------------------------------------------------------------------



from rs import *
from iqxel import *
#from CommonUtils import *
from common_utils import *
import numpy
import DUT_functions
from DUT_functions import *
from openpyxl import *

rw_num=1

op_file_path=""
ampl_dict={}
#Loop iterating for the range of amplitudes mentioned

NOMINAL_PER = 10 #Percentage
MARGIN = 2
LOW_PER_BOUNDARY = NOMINAL_PER - MARGIN
HIGH_PER_BOUNDARY = NOMINAL_PER + MARGIN

def get_10prcnt_per(ch,dr,start_amp,end_amp,step_size,min_per_val,max_per_val,a):
    global row_num
    global break_per
    global even_xaxis
    global odd_xaxis
    global data
    global row_num_dr
    steps_list=[]
    for i in numpy.arange(float(end_amp),float(start_amp),float(step_size)):
        steps_list.append(i)
    if(steps_list[0]!=float(end_amp)):
        steps_list.insert(0,float(end_amp))
    if(steps_list[-1]!=float(start_amp)):
        steps_list.append(float(start_amp))
    steps_list.reverse()
    dig_amplitude_list=[]
    for j in steps_list:
        dig_amplitude_list.append(j)
    global flag
    global limit_trail
    a=1
    if(limit_trail>7):
        #return float(curr_amp)-float(cable_loss)
        raise Exception(float(start_amp)-float(cable_loss))
    limit_trail=limit_trail+1
    for amp in dig_amplitude_list[1:]:
        if(flag==1):
            #return float(curr_amp)-float(cable_loss)
            raise Exception(float(curr_amp)-float(cable_loss))
        data=[]
        data.append(standard)
        data.append(int(ch))
        try:
            data.append(int(dr))
        except:
            data.append(dr)
        data.append(float(amp)-cable_loss)
        if(dig_amplitude_list.index(amp)==1):
            equipment.set_amplitude(streams,str(amp))
            data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment,'default')
        else:
            equipment.set_amplitude(streams,str(amp))
            data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment,'default')
        if(data[4] < 800 ):# and ((ampl_dict[dr].index(amp)==0) or (ampl_dict[dr].index(amp)==1))):
            time.sleep(1)
            dut.set_dut_channel(int(ch),p20_flag)
            for i in range(num_params_to_pop):
                data.pop()
            data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment,'default')
        DebugPrint(data)
        if(data[4]=='LOCKUP'):
            for i in range(num_params_to_pop):
                data.pop()
            dut.init_dut(standard=standard,streams=streams,channel=int(ch),bw=bw,release=release,chain_sel=chain_sel)
            dut.set_dut_production_mode_settings(standard=standard,ch=int(ch),bw=bw)
            dut.set_dut_channel(int(ch),p20_flag)
            data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment,'default')
            DebugPrint(data)
        if(a==1):
            prev_per=float(data[-1])
            curr_per=float(data[-1])
            a=a+1
            prev_amp=amp
            curr_amp=amp
        else:
            prev_per=max_per_val
            curr_per=float(data[-1])
            prev_amp=float(start_amp)
            curr_amp=amp
        if(float(curr_per)>HIGH_PER_BOUNDARY):
            worksheet_dr.write_row('A'+str(row_num_dr), data)
            row_num_dr=row_num_dr+1
            break_per=float(curr_per)
            get_10prcnt_per(ch,dr,float(start_amp),float(curr_amp),float(step_size/2),float(min_per_val),float(curr_per),a)
            if((break_per<float(curr_per))|(break_per>float(max_per_val))):
                break
        elif(LOW_PER_BOUNDARY<=float(curr_per)<=HIGH_PER_BOUNDARY):
            flag=1
            ten_perc_per_sig=str(curr_amp)
            sensitivity_amp_dict[ch][dr]=float(curr_amp)-float(cable_loss)
            worksheet.write_row('A'+str(row_num), data)
            worksheet_dr.write_row('A'+str(row_num_dr), data)
            if(flag==1):
                del(dig_amplitude_list[:])
            raise Exception( float(curr_amp)-float(cable_loss))
        elif(float(curr_per)==0):
            sensitivity_amp_dict[ch][dr]=float(curr_amp)-float(cable_loss)
            worksheet_dr.write_row('A'+str(row_num_dr), data)
            row_num_dr=row_num_dr+1
            break_per=float(curr_per)
            if((break_per<float(curr_per))|(break_per>float(max_per_val))):
                break
            get_10prcnt_per(ch,dr,float(curr_amp),float(end_amp),float(step_size/2),float(curr_per),float(max_per_val),a)
        elif(0<float(curr_per)<LOW_PER_BOUNDARY):
            sensitivity_amp_dict[ch][dr]=float(curr_amp)-float(cable_loss)
            worksheet_dr.write_row('A'+str(row_num_dr), data)
            row_num_dr=row_num_dr+1
            break_per=float(curr_per)
            if((break_per<float(curr_per))|(break_per>float(max_per_val))):
                break
            get_10prcnt_per(ch,dr,float(curr_amp),float(end_amp),float(step_size/2),float(curr_per),float(max_per_val),a)

def amplitude_loop(dr,ch,step_size):
    start_amp_list=[]
    global row_num
    global per_evenrow_dict
    per_evenrow_dict={}
    global per_oddrow_dict
    per_oddrow_dict={}
    global even_xaxis
    global odd_xaxis
    global data
    global row_num_dr
    row_num_dr=2


    #PHY_PERFORMANCE - work sheet heading contains lot of stats which are currently not available in calder.
    #worksheet.write_row('A1',['Standard','Channel','DataRate','AMPLITUDE','ED_CNT','CRC32_PASS_CNT','CRC32_FAIL_CNT','L1_CORR_FAIL_CNT','OFDM_S2L_FAIL_CNT','LSIG_FAIL_CNT','HTSIG_FAIL_CNT','VHTSIGA_FAIL_CNT','VHTSIGB_FAIL_CNT','DSSS_CORR_PASS_CNT','DSSS_SFD_FAIL_CNT','DSSS_HDR_FAIL_CNT','PER %'],bold1)
    #worksheet_dr.write_row('A1',['Standard','Channel','DataRate','AMPLITUDE','ED_CNT','CRC32_PASS_CNT','CRC32_FAIL_CNT','L1_CORR_FAIL_CNT','OFDM_S2L_FAIL_CNT','LSIG_FAIL_CNT','HTSIG_FAIL_CNT','VHTSIGA_FAIL_CNT','VHTSIGB_FAIL_CNT','DSSS_CORR_PASS_CNT','DSSS_SFD_FAIL_CNT','DSSS_HDR_FAIL_CNT','PER %'],bold)

    worksheet.write_row('A1',['Standard','Channel','DataRate','AMPLITUDE']+rx_report_hdngs,bold1)
    worksheet_dr.write_row('A1',['Standard','Channel','DataRate','AMPLITUDE']+rx_report_hdngs,bold)

    axis_num=1
    flg=0
    a=1
    global limit_trail
    global flag
    flag=0
    limit_trail=1
    res=dut.dut_read(1500)
    for amp in ampl_dict[dr]:
        print "--------------------------------------------------------VSG power level",amp  #PHY_PERFORMANCE
        data=[]
        data.append(standard)
        data.append(int(ch))
        try:
            data.append(int(dr))
        except:
            data.append(dr)
        data.append(amp-cable_loss)
        if(flg==0):
            if(ampl_dict[dr].index(amp)==0):
                equipment.set_amplitude(streams,str(amp))
                data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment,'default')
            else:
                equipment.set_amplitude(streams,str(amp))
                data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment,'default')
            #if((data[4] < 200 ) and ((ampl_dict[dr].index(amp)==0) or (ampl_dict[dr].index(amp)==1))):
            if(data[4] < 800 ):# and ((ampl_dict[dr].index(amp)==0) or (ampl_dict[dr].index(amp)==1))):
                time.sleep(1)
                dut.set_dut_channel(int(ch),p20_flag)
                #for i in range(13):
                for i in range(num_params_to_pop): #PHY_PERFORMANCE .. there are only 8 elements in data
                    data.pop()
                data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
            if((data[4] > 3000) and ('IQXEL' in tester)):
                time.sleep(2)
                #for i in range(13):
                for i in range(num_params_to_pop): #PHY_PERFORMANCE .. there are only 8 elements in data
                    data.pop()
                data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
            DebugPrint(data)
            if(data[4]=='LOCKUP'):
                #for i in range(13):
                for i in range(num_params_to_pop): #PHY_PERFORMANCE .. there are only 8 elements in data
                    data.pop()
                if(dut.check_dut_stuck_state()=='alive'):
                    dut.dut_reboot()
                else:
                    print 'DUT is stuck'
                    controlPowerSwitch(on_ports_list='8',off_ports_list='8')
                time.sleep(reboot_time)
                dut.init_dut(standard=standard,streams=streams,channel=int(ch),bw=bw,release=release,chain_sel=chain_sel)
                dut.set_dut_production_mode_settings(standard=standard,ch=int(ch),bw=bw)
                dut.set_dut_channel(int(ch),p20_flag)
                data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                DebugPrint(data)
            if(a==1):
                prev_per=data[-1]
                curr_per=data[-1]
                prev_amp=amp
                curr_amp=amp
            else:
                prev_per=curr_per
                curr_per=data[-1]
                prev_amp=curr_amp
                curr_amp=amp
            if(a==1):
                prev_data=list(data[4:])
                curr_data=list(data[4:])
            else:
                prev_data=list(curr_data)
                curr_data=list(data[4:])
            if((a!=1) and (not(cmp(prev_data,curr_data)))):
                DebugPrint('COMPARED')
                #for i in range(13):
                for i in range(num_params_to_pop): #PHY_PERFORMANCE .. there are only 8 elements in data
                    data.pop()
                data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                DebugPrint(data)
                if(data[4]=='LOCKUP'):
                #for i in range(13):
                    for i in range(num_params_to_pop): #PHY_PERFORMANCE .. there are only 8 elements in data
                        data.pop()
                    dut.dut_reboot()
                    time.sleep(reboot_time)
                    dut.init_dut(standard=standard,streams=streams,channel=int(channel.split(',')[-1]),bw=bw,release=release,chain_sel=chain_sel)
                    dut.set_dut_production_mode_settings(standard=standard,ch=int(ch),bw=bw)
                    dut.set_dut_channel(int(ch),p20_flag)
                    data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                    DebugPrint(data)
            a+=1
            if(flg==0):
                #worksheet.write_row('A'+str(row_num), data)
                worksheet_dr.write_row('A'+str(row_num_dr), data)
                #row_num=row_num+1
                row_num_dr=row_num_dr+1
            if(LOW_PER_BOUNDARY<=float(curr_per)<=HIGH_PER_BOUNDARY):
                #if(ampl_dict[dr].index(amp)>0):
                flg=1
                ten_perc_per_sig=str(curr_amp)
                sensitivity_amp_dict[ch][dr]=float(curr_amp)-float(cable_loss)
                worksheet.write_row('A'+str(row_num), data)
                #worksheet_dr.write_row('A'+str(row_num_dr), data)
                #row_num_dr=row_num_dr+1
                axis_num=axis_num+1
                ftime.write(time.strftime("%H-%M-%S")+"\n")
                return float(curr_amp)-float(cable_loss)
            elif(int(prev_per)<HIGH_PER_BOUNDARY<int(curr_per)):
                flg=1
                ten_perc_per_sig=str(prev_amp)+'-'+str(curr_amp)
                try:
                    sensity = get_10prcnt_per(ch,dr,float(prev_amp),float(curr_amp),float(step_size/2),float(prev_per),float(curr_per),1)
                except Exception,e:
                    print e.args
                    sensity = e.args[0]
            #elif(int(curr_per)>12):
            elif((int(curr_per)>HIGH_PER_BOUNDARY) or ((int(curr_per)<LOW_PER_BOUNDARY) and ((ampl_dict[dr].index(amp)==(len(ampl_dict[dr])-1))))):
                for amp in actual_ampl_dict[dr]:
                    data=[]
                    data.append(standard)
                    data.append(int(ch))
                    try:
                        data.append(int(dr))
                    except:
                        data.append(dr)
                    data.append(amp-cable_loss)
                    if(flg==0):
                        # if(actual_ampl_dict[dr].index(amp)==0):
                            # set_amplitude(str(int(amp)+1))
                            # time.sleep(5)
                        equipment.set_amplitude(streams,str(amp))
                        data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                        DebugPrint(data)
                        #if((data[4] < 1200 ) and ((actual_ampl_dict[dr].index(amp)==0) or (actual_ampl_dict[dr].index(amp)==1))):
                        if(data[4] < 800 ):# and ((actual_ampl_dict[dr].index(amp)==0) or (actual_ampl_dict[dr].index(amp)==1))):
                            time.sleep(8)
                            #for i in range(13):
                            for i in range(num_params_to_pop): #PHY_PERFORMANCE .. there are only 8 elements in data
                                data.pop()
                            data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                        if(data[4]=='LOCKUP'):
                            #for i in range(13):
                            for i in range(num_params_to_pop): #PHY_PERFORMANCE .. there are only 8 elements in data
                                data.pop()
                            dut.dut_reboot()
                            time.sleep(reboot_time)
                            dut.init_dut(standard=standard,streams=streams,channel=int(ch),bw=bw,release=release,chain_sel=chain_sel)
                            dut.set_dut_production_mode_settings(standard=standard,ch=int(ch),bw=bw)
                            dut.set_dut_channel(int(ch),p20_flag)
                            data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                            DebugPrint(data)
                        if(a==1):
                            prev_data=list(data[4:])
                            curr_data=list(data[4:])
                        else:
                            prev_data=curr_data
                            curr_data=list(data[4:])
                        if((a!=1) and (cmp(prev_data,curr_data)==0)):
                            #for i in range(13):
                            for i in range(num_params_to_pop): #PHY_PERFORMANCE .. there are only 8 elements in data
                                data.pop()
                            data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                            DebugPrint(data)
                            if(data[4]=='LOCKUP'):
                                #for i in range(13):
                                for i in range(num_params_to_pop): #PHY_PERFORMANCE .. there are only 8 elements in data
                                    data.pop()
                                dut.dut_reboot()
                                time.sleep(reboot_time)
                                dut.init_dut(standard=standard,streams=streams,channel=int(channel.split(',')[-1]),bw=bw,release=release,chain_sel=chain_sel)
                                dut.set_dut_production_mode_settings(standard=standard,ch=int(ch),bw=bw)
                                dut.set_dut_channel(int(ch),p20_flag)
                                data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                                DebugPrint(data)
                        if(a==1):
                            prev_per=data[-1]
                            curr_per=data[-1]
                            a=a+1
                            prev_amp=amp
                            curr_amp=amp
                        else:
                            prev_per=curr_per
                            curr_per=data[-1]
                            prev_amp=curr_amp
                            curr_amp=amp
                        if(flg==0):
                            #worksheet.write_row('A'+str(row_num), data)
                            worksheet_dr.write_row('A'+str(row_num_dr), data)
                            #row_num=row_num+1
                            row_num_dr=row_num_dr+1
                        if(LOW_PER_BOUNDARY<=float(curr_per)<=HIGH_PER_BOUNDARY):
                            flg=1
                            ten_perc_per_sig=str(curr_amp)
                            sensitivity_amp_dict[ch][dr]=float(curr_amp)-float(cable_loss)
                            worksheet.write_row('A'+str(row_num), data)
                            #worksheet_dr.write_row('A'+str(row_num_dr), data)
                            #row_num=row_num+1
                            #row_num_dr=row_num_dr+1
                            axis_num=axis_num+1
                            ftime.write(time.strftime("%H-%M-%S")+"\n")
                            return float(curr_amp)-float(cable_loss)
                        elif(int(prev_per)<HIGH_PER_BOUNDARY<int(curr_per)):
                            flg=1
                            ten_perc_per_sig=str(prev_amp)+'-'+str(curr_amp)
                            print '\nStarted to search 10% PER point'
                            try:
                                sensity = get_10prcnt_per(ch,dr,float(prev_amp),float(curr_amp),float(step_size/2),float(prev_per),float(curr_per),1)
                            except Exception,e:
                                print e.args
                                sensity = e.args[0]
                        elif((actual_ampl_dict[dr].index(amp)==0) and (int(curr_per)>HIGH_PER_BOUNDARY)):
                            return

    row_num=row_num+1
    return sensity

def readLookUpTable(sheet):
    """ Reading expected Sensitivity results from rxPerformanceReport.xlsx """
    reportXlsx = os.path.join(os.getcwd(),'../reports/530_77/rf/rxPerformanceReport.xlsx')
    workbook= load_workbook(reportXlsx)
    sheets = workbook.sheetnames
    if (sheet in sheets):
        ws = workbook[sheet]
        maxRow = ws.max_row
        colB = ws['B']
        colC = ws['F']
        sensitivityDict = {}
        for i in range(len(colB)-1):
            sensitivityDict.update({str(colB[i+1].value):colC[i+1].value})

        return sensitivityDict
    else:
        return 'NA'

def save_to_doc(hdng,freq,chn,bw):
    global document_path
    if (fadingtype != 'NA'):
        document_path=os.path.join(op_file_path.split(standard)[0],'Multipath_'+release+'_'+board_num+'_'+rf_num+'.docx')
    else:
        document_path=os.path.join(op_file_path.split(standard)[0],'RX_Characterization_Sensitivity_'+release+'_'+board_num+'_'+rf_num+'.docx')
    try:
        document = Document(document_path)
    except:
        document = Document()
    paragraphs_obj=document.paragraphs
    paragraphs=[]
    for p in paragraphs_obj:
        paragraphs.append(p.text)
    if (fadingtype != 'NA'):
        if(release+" RX Performance" not in paragraphs):
            document.add_heading(release+" RX Performance", level=1)
        if('Sensitivity of All DataRates' not in paragraphs):
            document.add_heading('Sensitivity of All DataRates', level=2)
        if(standard=='11ax'):
            if("HE - "+streams_chain+"-Channel-"+str(chn)+"-"+bw+"MHz" not in paragraphs):
                document.add_heading("HE - "+streams_chain+"-Channel-"+str(chn)+"-"+bw+"MHz"+'_fadeType_' + fadingtype, level=3)
        elif(standard=='11ac'):
            if("VHT - "+streams_chain+"-Channel-"+str(chn)+"-"+bw+"MHz" not in paragraphs):
                document.add_heading("VHT - "+streams_chain+"-Channel-"+str(chn)+"-"+bw+"MHz"+'_fadeType_' + fadingtype, level=3)
        elif(standard=='11n'):
            if("HT - "+streams_chain+"-"+freq+"-Channel-"+str(chn)+"-"+bw+"MHz" not in paragraphs):
                document.add_heading("HT - "+streams_chain+"-"+freq+"-Channel-"+str(chn)+"-"+bw+"MHz"+'_fadeType_' + fadingtype, level=3)
        elif(standard=='11g'):
            if("Legacy - Channel-"+str(chn) not in paragraphs):
                document.add_heading("Legacy - Channel-"+str(chn)+'_fadeType_' + fadingtype, level=3)
        elif(standard=='11a'):
            if("Legacy - Channel-"+str(chn) not in paragraphs):
                document.add_heading("Legacy - Channel-"+str(chn)+'_fadeType_' + fadingtype, level=3)
        elif(standard=='11b'):
            if("DSSS - Channel-"+str(chn) not in paragraphs):
                document.add_heading("DSSS - Channel-"+str(chn)+'_fadeType_' + fadingtype, level=3)
        if(hdng not in paragraphs):
            document.add_heading(hdng, level=4)

    else:
        if(release+" RX Performance" not in paragraphs):
            document.add_heading(release+" RX Performance", level=1)
        if('Sensitivity of All DataRates' not in paragraphs):
            document.add_heading('Sensitivity of All DataRates', level=2)
        if(standard=='11ax'):
            if("HE - "+streams_chain+"-Channel-"+str(chn)+"-"+bw+"MHz" not in paragraphs):
                document.add_heading("HE - "+streams_chain+"-Channel-"+str(chn)+"-"+bw+"MHz", level=3)
        elif(standard=='11ac'):
            if("VHT - "+streams_chain+"-Channel-"+str(chn)+"-"+bw+"MHz" not in paragraphs):
                document.add_heading("VHT - "+streams_chain+"-Channel-"+str(chn)+"-"+bw+"MHz", level=3)
        elif(standard=='11n'):
            if("HT - "+streams_chain+"-"+freq+"-Channel-"+str(chn)+"-"+bw+"MHz" not in paragraphs):
                document.add_heading("HT - "+streams_chain+"-"+freq+"-Channel-"+str(chn)+"-"+bw+"MHz", level=3)
        elif(standard=='11g'):
            if("Legacy - Channel-"+str(chn) not in paragraphs):
                document.add_heading("Legacy - Channel-"+str(chn), level=3)
        elif(standard=='11a'):
            if("Legacy - Channel-"+str(chn) not in paragraphs):
                document.add_heading("Legacy - Channel-"+str(chn), level=3)
        elif(standard=='11b'):
            if("DSSS - Channel-"+str(chn) not in paragraphs):
                document.add_heading("DSSS - Channel-"+str(chn), level=3)
        if(hdng not in paragraphs):
            document.add_heading(hdng, level=4)

    table = document.add_table(rows=1, cols=3)
    hdr_cells = table.rows[0].cells
    hdr_cells[0].text = 'DataRate'
    hdr_cells[1].text = 'Measured Sensitivity'
    hdr_cells[2].text = 'Expected sensitivity'
    if (standard=='11ax'):
        sensitivityDict = readLookUpTable(pktFrameFormat + '_' + bw )
    else:
        sensitivityDict = readLookUpTable(standard + '_' + bw )
    for datarate in data_rate.split(','):
        if(datarate in sensitivity_amp_dict[chn].keys()):
            row_cells = table.add_row().cells
            row_cells[0].text = str(datarate)
            if('TBT' in str(sensitivity_amp_dict[chn][datarate])):
                row_cells[1].text = str(sensitivity_amp_dict[chn][datarate])
            else:
                row_cells[1].text = str(round(float(sensitivity_amp_dict[chn][datarate]),1))
            if (sensitivityDict == 'NA'):
                row_cells[2].text = 'NA'
            else:
                row_cells[2].text = str(sensitivityDict[datarate])
    try:
        document.save(document_path)
    except:
        document.save(document_path+'_'+str(time.time())+'.docx')


#if __name__ == "__main__":
def sensitivity_analysis(DUT_TestConfigParams,dut1):
    #controlPowerSwitch(on_ports_list='8',off_ports_list='8')

    global ftime
    global equipment
    global dut
    global modulation
    global row_num
    global standard
    global streams_chain
    global streams
    global freq
    global bw
    global gi
    global stbc
    global coding
    global data_rate
    global cable_loss
    global greenfield_mode
    global channel
    global op_file_path
    global preamble
    global workbook
    global workbook_dr
    global worksheet
    global worksheet_dr
    global bold
    global bold1
    global actual_ampl_dict
    global tester
    global sensitivity_amp_dict
    global giType
    global heLtfType
    global pktFrameFormat
    global pltName
    global p20_flag
    global fadingtype

    greenfield_mode=stbc=preamble=gi=coding=preamble=''
    ftime=open('total_time','a')
    ftime.write(time.strftime("%H-%M-%S")+"\n")
    actual_ampl_dict={}
    sensitivity_amp_dict={}
    row_num=2
    dut = dut1

    # PHY_PERFORMANCE: is this the right place to call these?
##    DUT_functions.startNextCase() # PHY_PERFORMANCE - this was missing and is a bug
##    DUT_functions.startTestcase() # PHY_PERFORMANCE - this was missing and is a bug - without this DUT test wont start

    #standard=sys.argv[1].lower()
    #streams_chain=sys.argv[2]
    #streams=streams_chain.split('_')[0]
    #chain_sel=streams_chain.split('_')[-1]
    standard = DUT_TestConfigParams.standard #standard=sys.argv[1].lower()
    streams=DUT_TestConfigParams.streams        #streams=streams_chain.split('_')[0]
    chain_sel=DUT_TestConfigParams.chain_sel    #chain_sel=streams_chain.split('_')[-1]
    streams_chain = streams+'_'+str(chain_sel)
    bw=DUT_TestConfigParams.bw
    data_rate=DUT_TestConfigParams.data_rate
    dutModel=DUT_TestConfigParams.dutModel
    actual_start_amplt=DUT_TestConfigParams.start_amplt
    actual_end_amplt=DUT_TestConfigParams.end_amplt
    step_size=DUT_TestConfigParams.step_size
    channel=DUT_TestConfigParams.channel
    payload=DUT_TestConfigParams.payload
    p20_flag = DUT_TestConfigParams.p20_flag

    if (DUT_TestConfigParams.fadingtype != 'NA'):
        fadingtype = DUT_TestConfigParams.fadingtype
    else:
        fadingtype = 'NA'

    if(standard=='11b'):
        preamble=DUT_TestConfigParams.preamble

    else:
        stbc=DUT_TestConfigParams.stbc
        gi=DUT_TestConfigParams.gi
        coding=DUT_TestConfigParams.coding
        if(standard=='11n'):
            greenfield_mode=DUT_TestConfigParams.greenfield_mode
    if(standard=='11ax'):
        giType = str(0.8*(1<<DUT_TestConfigParams.giType))+'us'
        heLtfType = str((1<<DUT_TestConfigParams.heLtfType))+'x'
        tone = ''
        hemu_ru = ''
        if(DUT_TestConfigParams.format == 5):
            pktFrameFormat='HESU'
        elif(DUT_TestConfigParams.format == 6):
            pktFrameFormat='HEMU'
            ruAlloc = DUT_TestConfigParams.ruAllocation
            staId = DUT_TestConfigParams.desiredStaIDlist
            hemu_ru = 'RUAL_' + hex(ruAlloc) + '_STAID_' + staId + '_'
        elif(DUT_TestConfigParams.format == 7):
            pktFrameFormat='HEERSU'
            if (DUT_TestConfigParams.chBandwidth ==0):
                tone = '242Tone_'
            else:
                tone = '106Tone_'
        elif(DUT_TestConfigParams.format == 8):
            pktFrameFormat='HETB'
        dcm = 'DCM_'*DUT_TestConfigParams.dcm
        doppler = ('doppler'+str(DUT_TestConfigParams.midamblePeriodicity)+'_')*DUT_TestConfigParams.doppler
        pltName = pktFrameFormat+'_Chn'+str(channel)+'_'+bw+'Mhz_'+giType+'GI_'+heLtfType+'LTF_'+dcm+doppler+tone+hemu_ru+coding+'_'+stbc

    if(int(channel.split(',')[-1])<30):
        freq='2.4GHz'
    else:
        freq='5GHz'
    op_file_path=BuildResultsPath(DUT_TestConfigParams)
    time.sleep(0.5)
    if(standard=='11b'):
        modulation='DSSS'
    else:
        modulation='OFDM'
    ftime.write(time.strftime("%H-%M-%S")+"\n")
    DebugPrint(create='1')
    DebugPrint('Started')
    #dut=eval(dutModel)(com_port)
    tester = DUT_TestConfigParams.vsg
    equipment=eval(tester.split('_')[0])(tester,DUT_TestConfigParams.equip_ip)
    #dut.serial_access(standard=standard,streams=streams,channel=channel.split(',')[-1],bw=bw,release=release)
    DebugPrint('Trying to initialize dut in PER Main')
    if(dut.init_dut(standard=standard,streams=streams,channel=int(channel),bw=bw,release=release,chain_sel=chain_sel)=='CONNECT_FAIL'):
        print 'CONNECT_FAIL'
        DebugPrint('CONNECT_FAIL in PER Main')
        controlPowerSwitch(on_ports_list='8',off_ports_list='8')
        if(dut.init_dut(standard=standard,streams=streams,channel=int(channel.split(',')[0]),bw=bw,release=release,chain_sel=chain_sel)=='CONNECT_FAIL'):
            print 'Switching to Serial Access'
            DebugPrint('Switching to Serial Access')
            dutModel=dutModel.split('_')[0]
            dut=eval(dutModel)(com_port)
            equipment=eval(tester.split('_')[0])(tester,DUT_TestConfigParams.equip_ip)
            dut.init_dut(standard=standard,streams=streams,channel=int(channel.split(',')[0]),bw=bw,release=release,chain_sel=chain_sel)


    try:
        equipment.init_vsg_funcs(standard=standard,bw=bw,streams=streams,stbc=stbc,gi=gi,coding=coding,greenfield_mode=greenfield_mode,preamble=preamble,payload=payload,chain_sel=str(chain_sel))
        if(standard == '11ax'):
            equipment.configHE(DUT_TestConfigParams)
    except Exception,e:
        print 'Equipment is not reachable'
        exit(0)
    for dr in data_rate.split(','):
        amplitude_list=[]
        start_amplt=sensitivity_dict_range[standard][streams][bw.split('in')[0]][dr][0]
        end_amplt=sensitivity_dict_range[standard][streams][bw.split('in')[0]][dr][1]
        steps_list=list(numpy.arange(end_amplt,start_amplt+1,step_size))
        steps_list.reverse()
        for j in steps_list:
            amplitude_list.append(j)
        #if 0 not in amplitude_list:
        #    amplitude_list=[0]+amplitude_list
        #print "input power to be used for based on dataRate/MCS"
        print dr, amplitude_list #PHY_PERFORMANCE - added DR/MCS for better understanding
        ampl_dict[dr]=amplitude_list
        actual_amplitude_list=[]
        steps_list=list(numpy.arange(actual_end_amplt,actual_start_amplt+1,step_size))
        steps_list.reverse()
        for j in steps_list:
            actual_amplitude_list.append(j)
        actual_ampl_dict[dr]=actual_amplitude_list
    for ch in channel.split(','):
        if(int(ch)<30):
            freq='2.4GHz'
        else:
            freq='5GHz'
        if(1 <= int(ch) <= 14):
            cable_loss_band='24G_BAND'
        if(36 <= int(ch) <= 52):
            cable_loss_band='5G_BAND1'
        elif(53 <= int(ch) <= 108):
            cable_loss_band='5G_BAND2'
        elif(109 <= int(ch) <= 132):
            cable_loss_band='5G_BAND3'
        elif(133 <= int(ch) <= 165):
            cable_loss_band='5G_BAND4'
        cable_loss_1x1=cable_loss_dict['1x1'][cable_loss_band]
        cable_loss=cable_loss_2x2=cable_loss_1x1
        if(streams=='2x2'):
            cable_loss_2x2=cable_loss_dict['2x2'][cable_loss_band]
            cable_loss=float((cable_loss_1x1+cable_loss_2x2)/2)
        dut.set_dut_production_mode_settings(standard=standard,ch=int(ch),bw=bw)
        dut.set_dut_channel(int(ch),p20_flag)
        dut.dut_down_up(action='up_down')
        DUT_functions.startTestcase()
        equipment.apply_vsg(bw=bw,chn=int(ch),streams=streams,p20_flag=p20_flag)
        equipment.set_macheader()
        equipment.set_payload(standard,payload)
        if(standard=='11ac'):
            cnsl_fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch+'_'+bw+'Mhz_'+gi+'_'+coding+'_'+stbc
        elif(standard=='11ax'):
            cnsl_fname=pltName
            #cnsl_fname=bw+'MHz'
        elif(standard=='11n'):
            cnsl_fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch+'_'+bw+'Mhz_'+gi+'_'+coding+'_'+greenfield_mode+'_'+stbc
            #cnsl_fname=bw+'MHz'
        else:
            if(standard=='11b'):
                cnsl_fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch+'_'+preamble
            else:
                cnsl_fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch
            #cnsl_fname=''
        consld_fname=os.path.join(op_file_path,'Sensitivity_Consolidated_'+cnsl_fname+'.xlsx')
        workbook = xlsxwriter.Workbook(consld_fname)
        worksheet = workbook.add_worksheet()
        bold1 = workbook.add_format({'bold': 1})
        if(ch not in sensitivity_amp_dict.keys()):
            sensitivity_amp_dict[ch]={}
        if(standard=='11ax'):
            if (pktFrameFormat=='HEMU'):
                dut.dut_setStaID(int(DUT_TestConfigParams.desiredStaIDlist))
        for dr in data_rate.split(','):
            sensitivity_amp_dict[ch][dr]='TBT'
            if(str(dr).lower().find('mcs')>=0):
                if(standard=='11ac'):
                    fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch+'_'+bw+'Mhz_'+gi+'_'+coding+'_'+stbc+'_'+str(dr)
                elif(standard=='11ax'):
                    fname=pltName+'_'+str(dr)
                elif(standard=='11n'):
                    fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch+'_'+bw+'Mhz_'+gi+'_'+coding+'_'+'_'+greenfield_mode+'_'+stbc+'_'+str(dr)
            else:
                if(standard=='11b'):
                    fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch+'_'+preamble+'_'+str(dr)+'Mbps'
                else:
                    fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch+'_'+str(dr)+'Mbps'
            dr_file_name=os.path.join(op_file_path,'Sensitivity_'+fname+'.xlsx')
            workbook_dr = xlsxwriter.Workbook(dr_file_name,{'strings_to_urls': False})
            worksheet_dr = workbook_dr.add_worksheet()
            bold = workbook_dr.add_format({'bold': 1})
            equipment.set_datarate(modulation,standard,dr)
            # Setting datarate for all users in case of HEMU packet
            if(standard=='11ax'):
                if (pktFrameFormat=='HEMU'):
                    equipment.setDatarateAllUser(DUT_TestConfigParams.numUser, dr)

            equipment.generate_waveform(streams=streams)
            if (DUT_TestConfigParams.fadingtype != 'NA'):
                equipment.setFadingChannelType(DUT_TestConfigParams.fadingtype)
                equipment.enableFadingModule()
            ftime.write("AMP LOOP START"+time.strftime("%H-%M-%S")+"\n")
            senstivity = amplitude_loop(dr,ch,step_size)
            DebugPrint('sensititivity point is '+str(senstivity))
            ftime.write("AMP LOOP END"+time.strftime("%H-%M-%S")+"\n")
            try:
                workbook_dr.close()
            except:
                time.sleep(10)
                try:
                    workbook_dr.close()
                except:
                    pass
        try:
            save_to_doc(cnsl_fname,freq,ch,bw)
        except Exception,e:
            print e.args
            pass
    try:
        workbook.close()
    except:
        time.sleep(10)
    result_log_path = os.path.abspath('../Results/')
    result_log_path=os.path.join(result_log_path,DUT_TestConfigParams.release)
    result_log_path=os.path.join(result_log_path,'result_log_path_'+release+'.txt')
    f_log_path=open(result_log_path,'a')
    f_log_path.write('\nSENSITIVITY_'+cnsl_fname+'->'+op_file_path)
    f_log_path.close()
    equipment.rf_on_off(rf_state='off',streams=streams)
    if (DUT_TestConfigParams.fadingtype != 'NA'):
        equipment.disableFadingModule()
    equipment.close_vsg()
    dut.dut_close()
    return senstivity