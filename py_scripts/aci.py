#-------------------------------------------------------------------------------
# Name:        aci
# Purpose:     This script will calculate the aci value for the specified datarate.
# Author:      kranthi.kishore
# Created:     14-06-2015
# Copyright:   Imagination Technologies Ltd
#-------------------------------------------------------------------------------

from rs import *
from iqxel import *
from common_utils import *
from harness import *
from input import *
import numpy
from sensitivity import sensitivity_analysis
from openpyxl import *

op_file_path=""
rw_num=1
NOMINAL_PER = 10 #Percentage
MARGIN = 2
LOW_PER_BOUNDARY = NOMINAL_PER - MARGIN
HIGH_PER_BOUNDARY = NOMINAL_PER + MARGIN

op_file_path=""
sensitiviy_dr = {}

#Loop iterating for the range of amplitudes mentioned

def get_10prcnt_per(ch,dr,start_amp,end_amp,step_size,min_per_val,max_per_val,a):
    global row_num
    global break_per
    global even_xaxis
    global odd_xaxis
    global data
    global row_num_dr
    steps_list=[]
    for i in numpy.arange(float(start_amp),float(end_amp)+1,float(step_size)):
        steps_list.append(i)
    # if(steps_list[0]!=float(start_amp)):
        # steps_list.insert(0,float(start_amp))
    # if(steps_list[-1]!=float(end_amp)):
        # steps_list.append(float(end_amp))
    #steps_list.reverse()
    dig_amplitude_list=[]
    for j in steps_list:
        dig_amplitude_list.append(j)
    global flag
    global limit_trail
    a=1
    if(limit_trail>7):
        raise Exception('RAISED')
    limit_trail=limit_trail+1
    for amp in dig_amplitude_list[1:]:
        if(flag==1):
            raise Exception('RAISED')
        data=[]
        data.append(standard)
        data.append(int(ch))
        try:
            data.append(int(dr))
        except:
            data.append(dr)
        data.append(-1*((float(sensitiviy_dr[dr])+3)-(float(amp)-cable_loss)))
        data.append(float(sensitiviy_dr[dr])+3)
        data.append(float(amp)-cable_loss)
        if(dig_amplitude_list.index(amp)==1):
            equipment.set_aci_amplitude(streams,str(amp))
            data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
        else :
            equipment.set_aci_amplitude(streams,str(amp))
            data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
        print data
        if(data[6]=='LOCKUP'):
            for i in range(num_params_to_pop):
                data.pop()
            dut.init_dut(standard=standard,streams=streams,channel=int(ch),bw=bw,release=release,chain_sel=chain_sel)
            dut.set_dut_production_mode_settings(standard=standard,ch=int(ch),bw=bw)
            dut.set_dut_channel(int(ch))
            data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
            print data
        if(a==1):
            prev_per=float(data[-1])
            curr_per=float(data[-1])
            a=a+1
            prev_amp=amp
            curr_amp=amp
        else:
            prev_per=max_per_val
            curr_per=float(data[-1])
            prev_amp=float(start_amp)
            curr_amp=amp
        if(float(curr_per)>HIGH_PER_BOUNDARY):
            worksheet_dr.write_row('A'+str(row_num_dr), data)
            row_num_dr=row_num_dr+1
            break_per=float(curr_per)
            get_10prcnt_per(ch,dr,float(start_amp),float(curr_amp),float(step_size/2),float(min_per_val),float(curr_per),a)
            if((break_per<float(curr_per))|(break_per>float(max_per_val))):
                break
        elif(LOW_PER_BOUNDARY<=float(curr_per)<=HIGH_PER_BOUNDARY):
            flag=1
            ten_perc_per_sig=str(curr_amp)
            aci_amp_dict[ch][dr]=float(-1*((float(sensitiviy_dr[dr]))-(float(curr_amp)-cable_loss)))
            worksheet.write_row('A'+str(row_num), data)
            worksheet_dr.write_row('A'+str(row_num_dr), data)
            if(flag==1):
                del(dig_amplitude_list[:])
        elif(float(curr_per)==0):
            worksheet_dr.write_row('A'+str(row_num_dr), data)
            row_num_dr=row_num_dr+1
            break_per=float(curr_per)
            aci_amp_dict[ch][dr]=float(-1*((float(sensitiviy_dr[dr]))-(float(curr_amp)-cable_loss)))
            if((break_per<float(curr_per))|(break_per>float(max_per_val))):
                break
            get_10prcnt_per(ch,dr,float(curr_amp),float(end_amp),float(step_size/2),float(curr_per),float(max_per_val),a)
        elif(0<float(curr_per)<LOW_PER_BOUNDARY):
            worksheet_dr.write_row('A'+str(row_num_dr), data)
            row_num_dr=row_num_dr+1
            break_per=float(curr_per)
            aci_amp_dict[ch][dr]=float(-1*((float(sensitiviy_dr[dr]))-(float(curr_amp)-cable_loss)))
            if((break_per<float(curr_per))|(break_per>float(max_per_val))):
                break
            get_10prcnt_per(ch,dr,float(curr_amp),float(end_amp),float(step_size/2),float(curr_per),float(max_per_val),a)

def amplitude_loop(dr,ch):
    start_amp_list=[]
    global row_num
    global per_evenrow_dict
    per_evenrow_dict={}
    global per_oddrow_dict
    per_oddrow_dict={}
    global even_xaxis
    global odd_xaxis
    global data
    global row_num_dr
    row_num_dr=2
    worksheet.write_row('A1',['Standard','Channel','DataRate','Delta','Sensitivity + 3dB','Adjacent AMPLITUDE']+rx_report_hdngs,bold1)
    worksheet_dr.write_row('A1',['Standard','Channel','DataRate','Delta','Sensitivity + 3dB','Adjacent AMPLITUDE']+rx_report_hdngs,bold)

    axis_num=1
    flg=0
    a=1
    global limit_trail
    global flag
    flag=0
    limit_trail=1
    res=dut.dut_read(1500)
    for amp in ampl_dict[dr]:
        data=[]
        data.append(standard)
        data.append(int(ch))
        try:
            data.append(int(dr))
        except:
            data.append(dr)
        data.append(-1*((float(sensitiviy_dr[dr])+3)-(float(amp)-cable_loss)))
        data.append(float(sensitiviy_dr[dr])+3)
        data.append(float(amp)-cable_loss)
        if(flg==0):
            if(ampl_dict[dr].index(amp)==0):
                equipment.set_aci_amplitude(streams,str(int(amp)+1))
            equipment.set_aci_amplitude(streams,str(amp))
            data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
            #if((data[4] < 200 ) and ((ampl_dict[dr].index(amp)==0) or (ampl_dict[dr].index(amp)==1))):
            if(data[6] < 800 ):# and ((ampl_dict[dr].index(amp)==0) or (ampl_dict[dr].index(amp)==1))):
                time.sleep(1)
                dut.set_dut_channel(int(ch))
                for i in range(num_params_to_pop):
                    data.pop()
                data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
            if((data[6] > 3000) and ('IQXEL' in tester)):
                time.sleep(2)
                for i in range(num_params_to_pop):
                    data.pop()
                data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
            print data
            if(data[6]=='LOCKUP'):
                for i in range(num_params_to_pop):
                    data.pop()
                if(dut.check_dut_stuck_state()=='alive'):
                    dut.dut_reboot()
                else:
                    print 'DUT is stuck'
                    controlPowerSwitch(on_ports_list='8',off_ports_list='8')
                time.sleep(reboot_time)
                dut.init_dut(standard=standard,streams=streams,channel=int(ch),bw=bw,release=release,chain_sel=chain_sel)
                dut.set_dut_production_mode_settings(standard=standard,ch=int(ch),bw=bw)
                dut.set_dut_channel(int(ch))
                data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                print data
            if(a==1):
                prev_per=data[-1]
                curr_per=data[-1]
                prev_amp=amp
                curr_amp=amp
            else:
                prev_per=curr_per
                curr_per=data[-1]
                prev_amp=curr_amp
                curr_amp=amp
            if(a==1):
                prev_data=list(data[4:])
                curr_data=list(data[4:])
            else:
                prev_data=list(curr_data)
                curr_data=list(data[4:])
            if((a!=1) and (not(cmp(prev_data,curr_data)))):
                print 'COMPARED'
                for i in range(num_params_to_pop):
                    data.pop()
                data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                print data
                if(data[6]=='LOCKUP'):
                    for i in range(num_params_to_pop):
                        data.pop()
                    dut.dut_reboot()
                    time.sleep(reboot_time)
                    dut.init_dut(standard=standard,streams=streams,channel=int(channel.split(',')[-1]),bw=bw,release=release,chain_sel=chain_sel)
                    dut.set_dut_production_mode_settings(standard=standard,ch=int(ch),bw=bw)
                    dut.set_dut_channel(int(ch))
                    data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                    print data
            a+=1
            if(flg==0):
                #worksheet.write_row('A'+str(row_num), data)
                worksheet_dr.write_row('A'+str(row_num_dr), data)
                #row_num=row_num+1
                row_num_dr=row_num_dr+1
            if(LOW_PER_BOUNDARY<=float(curr_per)<=HIGH_PER_BOUNDARY):
                if(ampl_dict[dr].index(amp)>0):
                    flg=1
                    ten_perc_per_sig=str(curr_amp)
                    aci_amp_dict[ch][dr]=float(-1*((float(sensitiviy_dr[dr]))-(float(amp)-cable_loss)))
                    worksheet.write_row('A'+str(row_num), data)
                    worksheet_dr.write_row('A'+str(row_num_dr), data)
                    row_num_dr=row_num_dr+1
                    axis_num=axis_num+1
                    ftime.write(time.strftime("%H-%M-%S")+"\n")
            elif(int(prev_per)<HIGH_PER_BOUNDARY<int(curr_per)):
                flg=1
                ten_perc_per_sig=str(prev_amp)+'-'+str(curr_amp)
                try:
                    get_10prcnt_per(ch,dr,float(prev_amp),float(curr_amp),float(step_size/2),float(prev_per),float(curr_per),1)
                except Exception,e:
                    print e.args
            #elif(int(curr_per)>HIGH_PER_BOUNDARY):
            elif((int(curr_per)>HIGH_PER_BOUNDARY) or ((int(curr_per)<LOW_PER_BOUNDARY) and ((ampl_dict[dr].index(amp)==(len(ampl_dict[dr])-1))))):
                a=1
                #for i in range(num_params_to_pop):
                #        data.pop()
                #    data=data+start_per(modulation,dr,ch,streams,standard,tester)
                actual_ampl_list=actual_ampl_dict[dr]
                exit_flg=0
                print actual_ampl_list
                if(int(curr_per)<LOW_PER_BOUNDARY):
                    exit_flg=1
                    actual_ampl_list.reverse()
                print actual_ampl_list
                for amp in actual_ampl_list:
                    data=[]
                    data.append(standard)
                    data.append(int(ch))
                    try:
                        data.append(int(dr))
                    except:
                        data.append(dr)
                    data.append(-1*((float(sensitiviy_dr[dr])+3)-(float(amp)-cable_loss )))
                    data.append(float(sensitiviy_dr[dr])+3)
                    data.append(float(amp)-cable_loss)
                    if(flg==0):
                        # if(actual_ampl_dict[dr].index(amp)==0):
                            # set_aci_amplitude(str(int(amp)+1))
                            # time.sleep(5)
                        equipment.set_aci_amplitude(streams,str(amp))
                        data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                        print data
                        #if((data[4] < 1200 ) and ((actual_ampl_dict[dr].index(amp)==0) or (actual_ampl_dict[dr].index(amp)==1))):
                        if(data[6] < 800 ):# and ((actual_ampl_dict[dr].index(amp)==0) or (actual_ampl_dict[dr].index(amp)==1))):
                            time.sleep(8)
                            for i in range(num_params_to_pop):
                                data.pop()
                            data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                        if(data[6]=='LOCKUP'):
                            for i in range(num_params_to_pop):
                                data.pop()
                            dut.dut_reboot()
                            time.sleep(reboot_time)
                            dut.init_dut(standard=standard,streams=streams,channel=int(ch),bw=bw,release=release,chain_sel=chain_sel)
                            dut.set_dut_production_mode_settings(standard=standard,ch=int(ch),bw=bw)
                            dut.set_dut_channel(int(ch))
                            data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                            print data
                        if(a==1):
                            prev_data=list(data[4:])
                            curr_data=list(data[4:])
                        else:
                            prev_data=curr_data
                            curr_data=list(data[4:])
                        if((a!=1) and (cmp(prev_data,curr_data)==0)):
                            for i in range(num_params_to_pop):
                                data.pop()
                            data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                            print data
                            if(data[6]=='LOCKUP'):
                                for i in range(num_params_to_pop):
                                    data.pop()
                                dut.dut_reboot()
                                time.sleep(reboot_time)
                                dut.init_dut(standard=standard,streams=streams,channel=int(channel.split(',')[-1]),bw=bw,release=release,chain_sel=chain_sel)
                                dut.set_dut_production_mode_settings(standard=standard,ch=int(ch),bw=bw)
                                dut.set_dut_channel(int(ch))
                                data=data+start_per(modulation,dr,ch,streams,standard,tester,dut,equipment)
                                print data
                        if(a==1):
                            prev_per=data[-1]
                            curr_per=data[-1]
                            a=a+1
                            prev_amp=amp
                            curr_amp=amp
                        else:
                            prev_per=curr_per
                            curr_per=data[-1]
                            prev_amp=curr_amp
                            curr_amp=amp
                        if(flg==0):
                            #worksheet.write_row('A'+str(row_num), data)
                            worksheet_dr.write_row('A'+str(row_num_dr), data)
                            #row_num=row_num+1
                            row_num_dr=row_num_dr+1
                        if(LOW_PER_BOUNDARY<=float(curr_per)<=HIGH_PER_BOUNDARY):
                            flg=1
                            ten_perc_per_sig=str(curr_amp)
                            aci_amp_dict[ch][dr]=float(-1*((float(sensitiviy_dr[dr]))-(float(curr_amp)-cable_loss)))
                            worksheet.write_row('A'+str(row_num), data)
                            worksheet_dr.write_row('A'+str(row_num_dr), data)
                            #row_num=row_num+1
                            row_num_dr=row_num_dr+1
                            axis_num=axis_num+1
                            ftime.write(time.strftime("%H-%M-%S")+"\n")
                        elif(int(prev_per)<HIGH_PER_BOUNDARY<int(curr_per)):
                            flg=1
                            ten_perc_per_sig=str(prev_amp)+'-'+str(curr_amp)
                            print '\n++++++++++++++AMPLI LOOP++++++++++++++++++++++++++++++++++++++'
                            try:
                                get_10prcnt_per(ch,dr,float(prev_amp),float(curr_amp),float(step_size/2),float(prev_per),float(curr_per),1)
                            except Exception,e:
                                print e.args
                        # elif((actual_ampl_list.index(amp)==0) and (int(curr_per)>HIGH_PER_BOUNDARY)):
                            # if(exit_flg==1):
                                # return

    row_num=row_num+1


def save_to_doc(hdng,freq,chn,bw):
    global document_path
    document_path=os.path.join(op_file_path.split(standard)[0],'RX_Characterization_ACI_'+release+'.docx')
    try:
        document = Document(document_path)
    except:
        document = Document()
    paragraphs_obj=document.paragraphs
    paragraphs=[]
    for p in paragraphs_obj:
        paragraphs.append(p.text)
    if(release+" RX Performance" not in paragraphs):
        document.add_heading(release+" RX Performance", level=1)
    if('ACI of All DataRates' not in paragraphs):
        document.add_heading('ACI of All DataRates', level=2)
    if(standard=='11ax'):
        if("HE - Channel-"+str(chn) not in paragraphs):
            document.add_heading("HE - Channel-"+str(chn), level=3)
    if(standard=='11ac'):
        if("VHT - "+streams_chain+"-Channel-"+str(chn)+"-"+bw+"MHz" not in paragraphs):
            document.add_heading("VHT - "+streams_chain+"-Channel-"+str(chn)+"-"+bw+"MHz", level=3)
    if(standard=='11n'):
        if("HT - "+streams_chain+"-"+freq+"-Channel-"+str(chn)+"-"+bw+"MHz" not in paragraphs):
            document.add_heading("HT - "+streams_chain+"-"+freq+"-Channel-"+str(chn)+"-"+bw+"MHz", level=3)
    if(standard=='11g'):
        if("Legacy - Channel-"+str(chn) not in paragraphs):
            document.add_heading("Legacy - Channel-"+str(chn), level=3)
    if(standard=='11a'):
        if("Legacy - Channel-"+str(chn) not in paragraphs):
            document.add_heading("Legacy - Channel-"+str(chn), level=3)
    if(standard=='11b'):
        if("DSSS - Channel-"+str(chn) not in paragraphs):
            document.add_heading("DSSS - Channel-"+str(chn), level=3)
    if(hdng not in paragraphs):
        document.add_heading(hdng, level=5)
    table = document.add_table(rows=1, cols=2)
    table.style = 'TableGrid'
    hdr_cells = table.rows[0].cells
    col_names=['DataRate','Amplitude']
    hdr_cells = table.rows[0].cells
    for idx, name in enumerate(col_names):
        paragraph = hdr_cells[idx].paragraphs[0]
        run = paragraph.add_run(name)
        run.bold = True
    for datarate in data_rate.split(','):
        if(datarate in aci_amp_dict[chn].keys()):
            row_cells = table.add_row().cells
            row_cells[0].text = str(datarate)
            if('TBT' in str(aci_amp_dict[chn][datarate])):
                row_cells[1].text = str(aci_amp_dict[chn][datarate])
            else:
                row_cells[1].text = str(round(float(aci_amp_dict[chn][datarate]),2))
    try:
        document.save(document_path)
    except:
        document.save(document_path+'_'+str(time.time())+'.docx')

def readLookUpTable(sheet,datarate):
    """ Reading matlab SNR results from rxPerformanceReport.xlsx """
    reportXlsx = os.path.join(os.getcwd(),'../reports/530_77/rf/rxPerformanceReport.xlsx')
    wb= load_workbook(reportXlsx)
    sheets = wb.sheetnames
    if (sheet in sheets):
        ws = wb[sheet]
        rateDict = {'1':'2','2':'3','5.5':'4','11':'5',
            '6':'2','9':'3','12':'4','18':'5','24':'6','36':'7','48':'8','54':'9',
            'MCS0':'2','MCS1':'3','MCS2':'4','MCS3':'5','MCS4':'6','MCS5':'7','MCS6':'8','MCS7':'9',
            'MCS8':'10','MCS9':'11'}

        if (ws['G'+rateDict[datarate]].value == None) :
            print('sensitivity results not available')
            return None
        else:
            return ws['G'+rateDict[datarate]].value
    else:
        return None

def writeLookUpTable(sheet, datarate, sensitivity):
    """ writing Measured SNR and expected sensitivity results to rxPerformanceReport.xlsx """
    reportXlsx = os.path.join(os.getcwd(),'../reports/530_77/rf/rxPerformanceReport.xlsx')
    wb= load_workbook(reportXlsx)
    sheets = wb.sheetnames
    rateDict = {'1':'2','2':'3','5.5':'4','11':'5',
        '6':'2','9':'3','12':'4','18':'5','24':'6','36':'7','48':'8','54':'9',
        'MCS0':'2','MCS1':'3','MCS2':'4','MCS3':'5','MCS4':'6','MCS5':'7','MCS6':'8','MCS7':'9',
        'MCS8':'10','MCS9':'11'}

    if (sheet in sheets):
        ws = wb[sheet]
        ws['G'+rateDict[datarate]] = round(float(sensitivity),1)
    wb.save(reportXlsx)

##if __name__ == "__main__":
def aci_analysis(DUT_TestConfigParams,dut1):
    #controlPowerSwitch(on_ports_list='8',off_ports_list='8')
    global ftime
    global equipment
    global dut
    global modulation
    global row_num
    global standard
    global streams_chain
    global streams
    global freq
    global bw
    global gi
    global stbc
    global coding
    global data_rate
    global cable_loss
    global greenfield_mode
    global channel
    global op_file_path
    global preamble
    global workbook
    global workbook_dr
    global worksheet
    global worksheet_dr
    global bold
    global bold1
    global actual_ampl_dict
    global ampl_dict
    global aci_amp_dict
    global step_size
    greenfield_mode=stbc=preamble=gi=coding=preamble=''
    global ftime
    global tester
    global giType
    global heLtfType
    global pktFrameFormat
    global pltName
    global p20_flag

    ftime=open('total_time','a')
    ftime.write(time.strftime("%H-%M-%S")+"\n")
    ampl_dict={}
    actual_ampl_dict={}
    aci_amp_dict={}
    row_num=2
    dut = dut1

    standard = DUT_TestConfigParams.standard
    streams=DUT_TestConfigParams.streams
    chain_sel=DUT_TestConfigParams.chain_sel
    streams_chain = streams+'_'+str(chain_sel)
    bw=DUT_TestConfigParams.bw
    data_rate=DUT_TestConfigParams.data_rate
    dutModel=DUT_TestConfigParams.dutModel
    actual_start_amplt=DUT_TestConfigParams.start_amplt
    actual_end_amplt=DUT_TestConfigParams.end_amplt
    step_size=DUT_TestConfigParams.step_size
    channel=DUT_TestConfigParams.channel
    payload=DUT_TestConfigParams.payload
    p20_flag = DUT_TestConfigParams.p20_flag
    if(standard=='11b'):
        preamble=DUT_TestConfigParams.preamble

    elif((standard=='11ac')or(standard=='11n')):
        stbc=DUT_TestConfigParams.stbc
        gi=DUT_TestConfigParams.gi
        coding=DUT_TestConfigParams.coding
        if(standard=='11n'):
            greenfield_mode=DUT_TestConfigParams.greenfield_mode
    elif(standard=='11ax'):
        stbc=DUT_TestConfigParams.stbc
        gi=DUT_TestConfigParams.gi
        coding=DUT_TestConfigParams.coding
        giType = str(0.8*(1<<DUT_TestConfigParams.giType))+'us'
        heLtfType = str((1<<DUT_TestConfigParams.heLtfType))+'x'
        tone = ''
        hemu_ru = ''
        if(DUT_TestConfigParams.format == 5):
            pktFrameFormat='HESU'
        elif(DUT_TestConfigParams.format == 6):
            pktFrameFormat='HEMU'
            ruAlloc = DUT_TestConfigParams.ruAllocation
            staId = DUT_TestConfigParams.desiredStaIDlist
            hemu_ru = 'RUAL_' + str(ruAlloc) + '_STAID_' + staId + '_'
        elif(DUT_TestConfigParams.format == 7):
            pktFrameFormat='HEERSU'
            if (DUT_TestConfigParams.chBandwidth ==0):
                tone = '242Tone_'
            else:
                tone = '106Tone_'
        elif(DUT_TestConfigParams.format == 8):
            pktFrameFormat='HETB'
        dcm = 'DCM_'*DUT_TestConfigParams.dcm
        doppler = ('doppler'+str(DUT_TestConfigParams.midamblePeriodicity)+'_')*DUT_TestConfigParams.doppler
        pltName = pktFrameFormat+'_Chn'+str(channel)+'_'+bw+'Mhz_'+giType+'GI_'+heLtfType+'LTF_'+dcm+doppler+tone+hemu_ru+coding+'_'+stbc


    if(int(channel.split(',')[-1])<30):
        freq='2.4GHz'
    else:
        freq='5GHz'
    op_file_path=BuildResultsPath(DUT_TestConfigParams)
    time.sleep(0.5)
    if(standard=='11b'):
        modulation='DSSS'
    else:
        modulation='OFDM'
    ftime.write(time.strftime("%H-%M-%S")+"\n")
    DebugPrint(create='1')
    DebugPrint('Started')

    for dr in data_rate.split(','):
        amplitude_list=[]
        if(1 <= int(channel) <= 14):
            cable_loss_band='24G_BAND'
        if(36 <= int(channel) <= 52):
            cable_loss_band='5G_BAND1'
        elif(53 <= int(channel) <= 108):
            cable_loss_band='5G_BAND2'
        elif(109 <= int(channel) <= 132):
            cable_loss_band='5G_BAND3'
        elif(133 <= int(channel) <= 165):
            cable_loss_band='5G_BAND4'
        cable_loss=cable_loss_dict['1x1'][cable_loss_band]
        DUT_TestConfigParams.data_rate = dr
        sen = readLookUpTable(standard+'_'+bw,dr)
        if (sen == None):
            sensitivity = round(float(sensitivity_analysis(DUT_TestConfigParams,dut1)),1)
            sensitiviy_dr.update({dr:sensitivity})
            if (sensitivity < -40):
                writeLookUpTable(standard+'_'+bw, dr, sensitivity)
        else:
            sensitiviy_dr.update({dr:sen})
        #sensitiviy_dr.update({dr:-52.6625})
        start_amplt= sensitiviy_dr[dr]+3+cable_loss+ 50 #aaci_dict_range[standard][streams][bw][dr][0]
        end_amplt=start_amplt - 60#aci_dict_range[standard][streams][bw][dr][1]
        steps_list=list(numpy.arange(end_amplt,start_amplt+1,step_size))
        #steps_list.reverse()
        for j in steps_list:
            amplitude_list.append(j)
        #if 0 not in amplitude_list:
        #    amplitude_list=[0]+amplitude_list
        print amplitude_list
        ampl_dict[dr]=amplitude_list
        actual_amplitude_list=[]
        steps_list=list(numpy.arange(actual_end_amplt,actual_start_amplt+1,step_size))
        #steps_list.reverse()
        for j in steps_list:
            actual_amplitude_list.append(j)
        actual_ampl_dict[dr]=actual_amplitude_list

##    dut=eval(dutModel)(com_port)
    tester = DUT_TestConfigParams.vsg
    equipment=eval(tester.split('_')[0])(equip_ip = DUT_TestConfigParams.equip_ip)
    DebugPrint('Trying to initialize dut in PER Main')
    if(dut.init_dut(standard=standard,streams=streams,channel=int(channel.split(',')[0]),bw=bw,release=release,chain_sel=chain_sel)=='CONNECT_FAIL'):
        print 'CONNECT_FAIL'
        DebugPrint('CONNECT_FAIL in PER Main')
        controlPowerSwitch(on_ports_list='8',off_ports_list='8')
        if(dut.init_dut(standard=standard,streams=streams,channel=int(channel.split(',')[0]),bw=bw,release=release,chain_sel=chain_sel)=='CONNECT_FAIL'):
            print 'Switching to Serial Access'
            DebugPrint('Switching to Serial Access')
            dutModel=dutModel.split('_')[0]
            dut=eval(dutModel)(com_port)
            equipment=eval(tester.split('_')[0])(tester,DUT_TestConfigParams.equip_ip)
            dut.init_dut(standard=standard,streams=streams,channel=int(channel.split(',')[0]),bw=bw,release=release,chain_sel=chain_sel)


    dut.dut_down_up(action='up_down')
    DUT_functions.startTestcase()

    try:
        equipment.init_vsg_funcs(standard=standard,bw=bw,streams=streams,stbc=stbc,gi=gi,coding=coding,greenfield_mode=greenfield_mode,preamble=preamble,payload=payload,chain_sel=str(chain_sel))
        if(standard == '11ax'):
            equipment.configHE(DUT_TestConfigParams)
            print("configured HE")
    except Exception,e:
        print 'Equipment is not reachable'
        exit(0)
    equipment.set_aci_bandwidth(standard=standard,bw=bw)
    equipment.rf_on_off_aci()
    for ch in channel.split(','):
        if(int(ch)<30):
            freq='2.4GHz'
        else:
            freq='5GHz'
        if(1 <= int(ch) <= 14):
            cable_loss_band='24G_BAND'
        if(36 <= int(ch) <= 52):
            cable_loss_band='5G_BAND1'
        elif(53 <= int(ch) <= 108):
            cable_loss_band='5G_BAND2'
        elif(109 <= int(ch) <= 132):
            cable_loss_band='5G_BAND3'
        elif(133 <= int(ch) <= 165):
            cable_loss_band='5G_BAND4'
        cable_loss_1x1=cable_loss_dict['1x1'][cable_loss_band]
        cable_loss=cable_loss_2x2=cable_loss_1x1
        if(streams=='2x2'):
            cable_loss_2x2=cable_loss_dict['2x2'][cable_loss_band]
            cable_loss=float((cable_loss_1x1+cable_loss_2x2)/2)
        #cable_loss=cable_loss+splitter_loss
        dut.set_dut_production_mode_settings(standard=standard,ch=int(ch),bw=bw)
        dut.set_dut_channel(int(ch))
        equipment.apply_vsg(bw=bw,chn=int(ch),streams=streams)
        equipment.set_macheader()
        equipment.set_payload(standard,payload)
        equipment.apply_aci_vsg(bw=bw,chn=int(ch),streams=streams)
        equipment.set_aci_datarate(modulation,standard,dr)
        equipment.set_aci_macheader()
        equipment.set_aci_payload(standard,payload)
        equipment.set_aci_idleinterval(100)
        if(standard=='11ac'):
            cnsl_fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch+'_'+bw+'Mhz_'+gi+'_'+coding+'_'+stbc
            #cnsl_fname=bw+'MHz'
        elif(standard=='11ax'):
            cnsl_fname = pltName
        elif(standard=='11n'):
            cnsl_fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch+'_'+bw+'Mhz_'+gi+'_'+coding+'_'+greenfield_mode+'_'+stbc
            #cnsl_fname=bw+'MHz'
        else:
            if(standard=='11b'):
                cnsl_fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch+'_'+preamble
            else:
                cnsl_fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch
            #cnsl_fname=''
        consld_fname=os.path.join(op_file_path,'ACI_Consolidated_'+cnsl_fname+'.xlsx')
        workbook = xlsxwriter.Workbook(consld_fname)
        worksheet = workbook.add_worksheet()
        bold1 = workbook.add_format({'bold': 1})
        if(ch not in aci_amp_dict.keys()):
            aci_amp_dict[ch]={}
        for dr in data_rate.split(','):
            aci_amp_dict[ch][dr]='TBT'
            if(str(dr).lower().find('mcs')>=0):
                if(standard=='11ac'):
                    fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch+'_'+bw+'Mhz_'+gi+'_'+coding+'_'+stbc+'_'+str(dr)
                if(standard=='11ax'):
                    fname=pltName+'_'+str(dr)
                elif(standard=='11n'):
                    fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch+'_'+bw+'Mhz_'+gi+'_'+coding+'_'+greenfield_mode+'_'+stbc+'_'+str(dr)
            else:
                if(standard=='11b'):
                    fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch+'_'+preamble+'_'+str(dr)+'Mbps'
                else:
                    fname=standard+'_'+streams_chain+'_'+freq+'_Chn'+ch+'_'+str(dr)+'Mbps'
            dr_file_name=op_file_path+'\ACI'+fname+'.xlsx'
            workbook_dr = xlsxwriter.Workbook(dr_file_name,{'strings_to_urls': False})
            worksheet_dr = workbook_dr.add_worksheet()
            bold = workbook_dr.add_format({'bold': 1})
            equipment.set_datarate(modulation,standard,dr)
            equipment.set_aci_datarate(modulation,standard,dr)
            equipment.set_idleinterval(250,chain_sel=chain_sel)
            equipment.set_aci_idleinterval(1)
            equipment.generate_waveform(streams=streams)
            equipment.generate_aci_waveform(streams=streams)
            ftime.write("AMP LOOP START"+time.strftime("%H-%M-%S")+"\n")
            if (DUT_TestConfigParams.in_band_pkt == 1):
                if(standard=='11b'):
                    equipment.set_amplitude(streams,str(float(sensitiviy_dr[dr]+3+cable_loss)))#TODO Adjustment AS per Sensitivity Values with Splitter, +10 is given for debugging remove later
                else:
                    equipment.set_amplitude(streams,str(float(sensitiviy_dr[dr]+3+cable_loss)))#TODO Adjustment AS per Sensitivity Values with Splitter, +10 is given for debugging remove later
            else:
                equipment.set_amplitude(streams,'-100.0')

            amplitude_loop(dr,ch)
            ftime.write("AMP LOOP END"+time.strftime("%H-%M-%S")+"\n")
            try:
                workbook_dr.close()
            except:
                time.sleep(10)
                try:
                    workbook_dr.close()
                except:
                    pass
            #copy_file(op_dr_file_path=dr_file_name,dutModel=dutModel,release=release,standard=standard,streams=streams_chain,stbc=stbc,coding=coding,gi=gi,greenfield_mode=greenfield_mode,preamble=preamble,ch=ch,test='rx')
        try:
            save_to_doc(cnsl_fname,freq,ch,bw)
        except Exception,e:
            print e.args
            pass
    try:
        workbook.close()
    except:
        time.sleep(10)
    #copy_file(op_dr_file_path=consld_fname,dutModel=dutModel,release=release,standard=standard,streams=streams_chain,stbc=stbc,coding=coding,gi=gi,greenfield_mode=greenfield_mode,preamble=preamble,ch=ch,test='rx')
    result_log_path = os.path.abspath('../Results/')
    result_log_path=os.path.join(result_log_path,DUT_TestConfigParams.release)
    result_log_path=os.path.join(result_log_path,'result_log_path_'+release+'.txt')
    f_log_path=open(result_log_path,'a')
    f_log_path.write('\nACI_'+cnsl_fname+'->'+op_file_path)
    f_log_path.close()
    equipment.rf_on_off(rf_state='off',streams=streams)
    equipment.rf_on_off_aci(rf_state='off',streams=streams)
    equipment.close_vsg()
    #copy_file(op_dr_file_path=document_path,dutModel=dutModel,release=release)
    dut.dut_close()