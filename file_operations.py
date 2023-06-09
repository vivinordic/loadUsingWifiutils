##################################################
import sys
import subprocess
import os
import openpyxl
from ctypes import *
from array import *
import xlrd

from lmac_def import *
##################################################

def readonce():
    #workbook = xlrd.open_workbook(test_config_xlsx)
    # Read the common Params from the workbook
    #commonParamsSheet = workbook.sheet_by_name('TestCases')

    test_config_xlsx="testPlans1.0.xlsx"
    global workbook
    workbook=openpyxl.load_workbook(test_config_xlsx , data_only=True)
    
    global constParamsSheet
    #ConstWorkbook =openpyxl.load_workbook(test_config_xlsx)
    # Read the common Params from the workbook
    constParamsSheet = workbook['InitializationParams']

    rowCount_constParams=constParamsSheet.max_row
    colCount_constParams=constParamsSheet.max_column
    
    global sheetName
    sheetName=constParamsSheet.cell(row=8,column=2).value
    print("sheets to run are " +str(sheetName))
    
    
def setofSheets():
    global sheetSet
    sheetSet=workbook.sheetnames
    return sheetSet

def commonParamsInfo(sheet):
    print("sheet to use is " +str(sheet))
    global commonParamsSheet
    commonParamsSheet=workbook[sheet]
    global rowCount_commonParams
    global colCount_commonParams
    rowCount_commonParams=commonParamsSheet.max_row
    #print("number of rows are " +str(rowCount_commonParams))
    colCount_commonParams=commonParamsSheet.max_column
    #print("number of columns are" +str(colCount_commonParams))
    sheetinfo=[]
    sheetinfo.append(rowCount_commonParams)
    sheetinfo.append(colCount_commonParams)
    return sheetinfo

def accessFile(rowNumber):
    global paramsDict
    #since readonce is already called during getTargetName function, no need to call again
    commonParamsCells = commonParamsSheet[rowNumber]
    #iter_rows
    #commonParamsCells = commonParamsSheet.row_slice(rowNumber, start_colx=0, end_colx=None)#reading from row 1

    paramsList = [] ; paramsDict = {}
    #header =commonParamsSheet.row_values(0, start_colx=0, end_colx=None)

    print('row number is' +str(rowNumber))
    for cell in commonParamsCells:
        paramsList.append(cell.value)
        #print(cell.value) #uncomment this line to see what values are read from excel sheet
    index = 0
    #print(WifiStandard(1));#print(WifiStandard[paramsList[index]].value)
    paramsDict.update( {'TestName'          : TestId[paramsList[index]].value} ) ; index += 1
    paramsDict.update( {'FrameSource'       : paramsList[index]} ) ; index += 1
    paramsDict.update( {'ProtectionType'    : paramsList[index]} ) ; index += 1
    paramsDict.update( {'TestType'          : WifiStandard[paramsList[index]].value} ) ; index += 1
    paramsDict.update( {'SecurityMode'      : SecurityModes[paramsList[index]].value} ) ; index += 1
    paramsDict.update( {'TxPacketNumber'    : paramsList[index]} ) ; index += 1
    paramsDict.update( {'AmpduSubFrameCount': paramsList[index]} ) ; index += 1
    paramsDict.update( {'PacketType'        : SelectValue[paramsList[index]].value} ) ; index += 1
    #paramsDict.update( {'PacketSize'        : paramsList[index]} ) ; index += 1
    paramsDict.update( {'PacketSize'        : 100} ) ; index += 1
    paramsDict.update( {'DataRate'          : DataRate[paramsList[index]].value} ) ; index += 1
    paramsDict.update( {'Aggregation'       : SetValue[paramsList[index]].value} ) ; index += 1
    paramsDict.update( {'LDPC'              : SetValue[paramsList[index]].value} ) ; index += 1
    paramsDict.update( {'STBC'              : SetValue[paramsList[index]].value} ) ; index += 1
    paramsDict.update( {'ChannelBW'         : ChannelSize[paramsList[index]].value} ) ; index += 1
    paramsDict.update( {'SIG'               : PacketSGI[paramsList[index]].value} ) ; index += 1
    paramsDict.update( {'PacketFormat'      : PacketFormat[paramsList[index]].value} ) ; index += 1
    paramsDict.update( {'DSSSPreamble'      : DsssPreamble[paramsList[index]].value} ) ; index += 1
    paramsDict.update( {'OFDMPreamble'      : OfdmPreamble[paramsList[index]].value} ) ; index += 1
    if (paramsDict["DataRate"] in [2, 4, 11, 22, 12, 18, 24, 36, 48, 72, 96, 108]) :
        paramsDict["Aggregation"]=0
    if (paramsDict["Aggregation"]==0):
        paramsDict["AmpduSubFrameCount"]=1
    if (paramsDict["SecurityMode"]!=SecurityModes.RPU_PLAIN.value):
        paramsDict.update( {'QoSParam'          : paramsList[index]} ) ; index += 1
        paramsDict.update( {'BroadcastParam'    : paramsList[index]} ) ; index +=1
        paramsDict.update( {'A4AddrParam'       : paramsList[index]} ) ; index +=1
        paramsDict.update( {'KeyId'             : paramsList[index]} ) ; index +=1
        paramsDict.update( {'DBNum'             : paramsList[index]} ) ; index +=1
    print(paramsDict)
    return paramsDict


def constantParams():
    #constParamsCells = constParamsSheet.col_slice(1, start_rowx=0, end_rowx=None)#reading from column 1
    #constParamsCells = constParamsSheet['B']
    constParamsList = [] ; constParamsDict = {}

    #print('value in file operations')
    #for cell in constParamsCells:
        #constParamsList.append(cell.value)
        #print(cell.value)

    #row and column numbering begins at 1
    constParamsDict.update( {'SheetName'           : constParamsSheet.cell(row=8,column=2).value} ) ;  # row 7, col 2
    constParamsDict['MacAddress']={}
    for i in range(0,6):
        constParamsDict['MacAddress'][i]= constParamsSheet.cell(row=9,column=i+2).value
    constParamsDict.update( {'ChannelNumber'        : constParamsSheet.cell(row=10,column=2).value} ) ;
    print(constParamsDict)
    return constParamsDict

def getTargetName():
    readonce()
    targetParams= {}
    targetParams.update({'TargetName' : constParamsSheet.cell(row=1,column=2).value})
    #print(targetParams['TargetName'])
    targetParams.update({'SimNumber' : constParamsSheet.cell(row=2,column=2).value})
    targetParams.update({'FpgaNumber' : constParamsSheet.cell(row=3,column=2).value})
    targetParams.update({'EmuNumber' : constParamsSheet.cell(row=6,column=2).value})
    targetParams.update({'elfLocation' : constParamsSheet.cell(row=7,column=2).value})
    if (targetParams['TargetName'] == 'FPGA'):
        targetParams.update({'FPGALocation' : constParamsSheet.cell(row=4,column=2).value})
    if (targetParams['TargetName'] == 'SILICON'):
        targetParams.update({'SiliconAccessType' : constParamsSheet.cell(row=5, column=2).value})
        targetParams.update({'SiliconLocation' : constParamsSheet.cell(row=4, column=2).value})
        """ for now silicon location set to FPGA location - change later """
    return targetParams

def writeToFile(result, rowNum):
    cell=commonParamsSheet.cell(row=rowNum, column = colCount_commonParams+1)
    if(result==1):
        print("PASS")
        cell.value="PASS"
        #commonParamsSheet.write(rowNum, colCount_commonParams+1, 'PASS')
    elif(result==0):
        print("FAIL")
        cell.value="FAIL"
        #commonParamsSheet.write(rowNum, colCount_commonParams+1, 'FAIL')
    #workbook.save(filename="testPlans1.0.xlsx")


def reference():
    #setting the variable type to pointer to structure
    paramsPtr=POINTER(LMAC_TB_PARAMS)

    #setting variable type to structure
    #argParams=LMAC_TB_PARAMS()

    #declaring intance of the class params
    #testParams=params()

    #reading params from excel sheet
    #params.frameSrc=int(commonParamsSheet.cell_value(1,1))
    #params.protectionType=0
    #params.tx_pkt_num=int(commonParamsSheet.cell_value(1,3))
    #params.ampdu_subframe_cnt=int(commonParamsSheet.cell_value(1,4))

    #displaying params list from excel sheet
    #paramsList=[]
    #paramsList=file_operations.accessFile()

    #for i in range(len(paramsList)):
    #    print(paramsList[i])

    #lmacTb.wlanSendTxPkt.argtypes = [paramsPtr]

    #print( 'the value of cmd buf is ' +str(TxCmd))
    #lmacMacAddrConfCmd=c_char_p.from_buffer(lmacMacAddrConfCmdPtr)

    #for i in TxCmd:
        #print(hex(i))
    #print("Bytes in Python {0}".format(lmacMacAddrConfCmd.value))
    #macAddConfList=[]
    #macAddConfList[:0]=lmacMacAddrConfCmd
    #print(macAddConfList)